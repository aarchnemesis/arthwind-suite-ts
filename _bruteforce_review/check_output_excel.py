import openpyxl

path = r"C:\Users\Pedro\Downloads\Final\VSR-05-04\VSR-05-04_Novo_Excel.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.worksheets[0]

headers = []
for c in ws[1]:
    headers.append(c.value)
print("Headers:", headers)

df_start_col = headers.index("DF distance - Start (m)") + 1
df_end_col = headers.index("DF distance - End (m)") + 1

print(f"\nDF Start col={df_start_col}, DF End col={df_end_col}\n")
for r in range(2, min(ws.max_row, 15) + 1):
    cs = ws.cell(row=r, column=df_start_col)
    ce = ws.cell(row=r, column=df_end_col)
    print(f"row {r}: start value={cs.value!r} type={type(cs.value).__name__} data_type={cs.data_type} numFmt={cs.number_format}  |  "
          f"end value={ce.value!r} type={type(ce.value).__name__} data_type={ce.data_type} numFmt={ce.number_format}")
