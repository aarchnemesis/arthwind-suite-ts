import json
import openpyxl
import os

EXCEL_FILES = {
    "VSR05-06": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR05-06.xlsx",
    "VSR07-04": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR07-04.xlsx",
}

# Build defectId -> inspection date map
date_by_id = {}
for turbine, path in EXCEL_FILES.items():
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column
    id_col = headers.get("defect id")
    date_col = headers.get("inspection date")
    for r in range(2, ws.max_row + 1):
        did = ws.cell(row=r, column=id_col).value
        date = ws.cell(row=r, column=date_col).value
        if did:
            date_by_id[str(did)] = str(date)

with open("_combined_calibration.json") as f:
    calib = json.load(f)

by_date = {}
for r in calib:
    date = date_by_id.get(r["defectId"], "UNKNOWN")
    day = date.split(" ")[0:3]  # crude
    by_date.setdefault(date[:16], []).append(r)

print("Distinct inspection dates in the calibration set:")
for date, rows in sorted(by_date.items()):
    scale_hs = [r["scale_h"] for r in rows if r["scale_h"]]
    avg = sum(scale_hs) / len(scale_hs) if scale_hs else None
    print(f"  {date}: {len(rows)} defects, avg scale_h={avg}")
