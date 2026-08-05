"""
Fits the equirect->pinhole camera projection (ported from Arthnex's own
resolveCamera/toCameraRay/clipToFrame) against the 85-point calibration set,
searching for the srcWidth/srcHeight that best explains the measured ground truth.
"""
import json
import math
import numpy as np
import openpyxl

PINHOLE_W, PINHOLE_H = 5568, 4176
ASPECT = PINHOLE_W / PINHOLE_H
MIN_HFOV_DEG, MAX_HFOV_DEG = 90, 140
FRAME_MARGIN = 1.1
MIN_RAY_Z = math.cos(math.radians(80))

EXCEL_FILES = {
    "VSR05-06": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR05-06.xlsx",
    "VSR07-04": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR07-04.xlsx",
    "VSR22-02": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR22-02.xlsx",
}


def load_points_by_id():
    points_by_id = {}
    for turbine, path in EXCEL_FILES.items():
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column
        id_col = headers.get("defect id")
        poly_col = headers.get("polygon data/coordinates")
        for r in range(2, ws.max_row + 1):
            did = ws.cell(row=r, column=id_col).value
            poly_raw = ws.cell(row=r, column=poly_col).value
            if not did or not poly_raw:
                continue
            try:
                pts = json.loads(poly_raw)
            except Exception:
                continue
            points_by_id[str(did)] = [{"x": p["x"], "y": abs(p["y"])} for p in pts]
    return points_by_id


def equirect_polygon_center(points, pano_width):
    y = sum(p["y"] for p in points) / len(points)
    reference = points[0]["x"]
    half = pano_width / 2
    offset = 0
    for p in points:
        dx = p["x"] - reference
        if dx > half:
            dx -= pano_width
        elif dx < -half:
            dx += pano_width
        offset += dx
    offset /= len(points)
    x = ((reference + offset) % pano_width + pano_width) % pano_width
    return x, y


def to_camera_ray(point, src_w, src_h, yaw, pitch):
    theta = ((point["x"] / src_w) * 2 - 1) * math.pi
    phi = (0.5 - point["y"] / src_h) * math.pi
    world_x = math.cos(phi) * math.sin(theta)
    world_y = math.sin(phi)
    world_z = math.cos(phi) * math.cos(theta)
    cos_yaw, sin_yaw = math.cos(yaw), math.sin(yaw)
    cos_pitch, sin_pitch = math.cos(pitch), math.sin(pitch)
    x = world_x * cos_yaw - world_z * sin_yaw
    z = world_x * sin_yaw + world_z * cos_yaw
    y = world_y * cos_pitch + z * sin_pitch
    zz = -world_y * sin_pitch + z * cos_pitch
    return x, y, zz


def resolve_camera(points, src_w, src_h):
    cx, cy = equirect_polygon_center(points, src_w)
    yaw = ((cx / src_w) * 2 - 1) * math.pi
    pitch = (cy / src_h - 0.5) * math.pi

    hfov_deg = MIN_HFOV_DEG
    for p in points:
        rx, ry, rz = to_camera_ray(p, src_w, src_h, yaw, pitch)
        if rz <= MIN_RAY_Z:
            hfov_deg = MAX_HFOV_DEG
            break
        hfov_deg = max(
            hfov_deg,
            math.degrees(2 * math.atan((abs(rx) / rz) * FRAME_MARGIN)),
            math.degrees(2 * math.atan((abs(ry) / rz) * FRAME_MARGIN)) * ASPECT,
        )
    hfov_deg = min(hfov_deg, MAX_HFOV_DEG)
    focal_x = PINHOLE_W / (2 * math.tan(math.radians(hfov_deg) / 2))
    focal_y = PINHOLE_H / (2 * math.tan(math.radians(hfov_deg / ASPECT) / 2))
    return yaw, pitch, focal_x, focal_y


def project_polygon_center(points, src_w, src_h):
    """Returns the pinhole-space centroid of the projected polygon (no frame clipping,
    just the raw projected points averaged — good enough for center-error scoring)."""
    yaw, pitch, focal_x, focal_y = resolve_camera(points, src_w, src_h)
    proj = []
    for p in points:
        rx, ry, rz = to_camera_ray(p, src_w, src_h, yaw, pitch)
        if rz <= MIN_RAY_Z:
            rz = MIN_RAY_Z  # avoid blowup; crude clip
        px = (rx / rz) * focal_x + PINHOLE_W / 2
        py = PINHOLE_H / 2 - (ry / rz) * focal_y
        proj.append((px, py))
    xs = [p[0] for p in proj]
    ys = [p[1] for p in proj]
    return sum(xs) / len(xs), sum(ys) / len(ys), max(xs) - min(xs), max(ys) - min(ys)


def main():
    with open("_all3_calibration.json") as f:
        calib = json.load(f)
    points_by_id = load_points_by_id()

    dataset = []
    for r in calib:
        pts = points_by_id.get(r["defectId"])
        if not pts or len(pts) < 3:
            continue
        dataset.append((r, pts))
    print(f"Dataset size: {len(dataset)}")

    def score(src_w, src_h):
        errs = []
        for r, pts in dataset:
            try:
                pcx, pcy, pw, ph = project_polygon_center(pts, src_w, src_h)
            except Exception:
                continue
            err = math.hypot(pcx - r["measured_cx"], pcy - r["measured_cy"])
            errs.append(err)
        return np.mean(errs), np.median(errs), len(errs)

    print("\n=== Coarse grid search over srcWidth (srcHeight = srcWidth assumed square-ish equirect via /2) ===")
    best = None
    for src_w in range(2000, 12000, 500):
        src_h = src_w / 2  # assume classic 2:1 equirect
        mean_err, median_err, n = score(src_w, src_h)
        marker = ""
        if best is None or median_err < best[2]:
            best = (src_w, src_h, median_err)
            marker = "  <-- best so far"
        print(f"  srcW={src_w:6d} srcH={src_h:7.1f}  mean_err={mean_err:8.1f} median_err={median_err:8.1f}{marker}")

    print(f"\nBest coarse: srcW={best[0]} srcH={best[1]:.1f} median_err={best[2]:.1f}")

    print("\n=== Fine search around best, also varying srcH independently ===")
    best_fine = None
    for src_w in range(max(1000, best[0] - 1000), best[0] + 1000, 100):
        for src_h in range(max(500, int(best[1] - 800)), int(best[1] + 800), 100):
            mean_err, median_err, n = score(src_w, src_h)
            if best_fine is None or median_err < best_fine[2]:
                best_fine = (src_w, src_h, median_err, mean_err)

    print(f"Best fine: srcW={best_fine[0]} srcH={best_fine[1]} median_err={best_fine[2]:.1f} mean_err={best_fine[3]:.1f}")

    # also test literal native 5568x4176 and 5568x2784(2:1) for reference
    for label, (sw, sh) in {
        "native 5568x4176": (5568, 4176),
        "native 5568x2784 (2:1)": (5568, 2784),
        "4K 3840x1920": (3840, 1920),
        "5.7K 5760x2880": (5760, 2880),
    }.items():
        mean_err, median_err, n = score(sw, sh)
        print(f"  reference [{label}]: median_err={median_err:.1f} mean_err={mean_err:.1f}")


if __name__ == "__main__":
    main()
