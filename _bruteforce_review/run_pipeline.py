"""
Pipeline reutilizável: extrai as imagens marcadas do PDF do relatório Arthnex,
casa cada uma com o Defect ID, mede a posição real do polígono por cor (severidade),
e cruza com a coordenada crua da planilha — gerando um dataset de calibração.

Uso:
    python run_pipeline.py "<caminho do excel>" "<caminho do pdf>" "<pasta de saída>"
"""
import fitz
import re
import json
import os
import sys
import openpyxl
import numpy as np
from PIL import Image
from scipy import ndimage

DEFECT_ID_RE = re.compile(r'^[A-Z0-9]{8,16}$')
SEVERITY_COLORS = {
    1: np.array([15, 181, 90]),    # green
    2: np.array([255, 192, 0]),    # gold
    3: np.array([204, 102, 0]),    # dark orange
}
COLOR_TOLERANCE = 25
NATIVE_W, NATIVE_H = 5568, 4176


def extract_pdf_images(pdf_path, out_dir):
    doc = fitz.open(pdf_path)
    manifest = []
    seen_ids = set()

    for pno in range(doc.page_count):
        page = doc[pno]
        text = page.get_text()
        lines = [l.strip() for l in text.split('\n') if l.strip()]

        defect_ids_on_page = []
        for i, line in enumerate(lines):
            if line == "Description" and i + 1 < len(lines):
                candidate = lines[i + 1]
                if DEFECT_ID_RE.match(candidate):
                    defect_ids_on_page.append(candidate)

        if not defect_ids_on_page:
            continue

        images = page.get_images(full=True)
        photo_images = [(img[0], img[2], img[3]) for img in images if img[2] >= 1500 and img[3] >= 1000]

        if len(photo_images) != len(defect_ids_on_page):
            print(f"[WARN] page {pno+1}: {len(defect_ids_on_page)} defect IDs but {len(photo_images)} photo-like images")

        for idx, defect_id in enumerate(defect_ids_on_page):
            if defect_id in seen_ids or idx >= len(photo_images):
                continue
            seen_ids.add(defect_id)
            xref, w, h = photo_images[idx]
            base = doc.extract_image(xref)
            fname = f"{defect_id}.{base['ext']}"
            with open(os.path.join(out_dir, fname), "wb") as f:
                f.write(base["image"])
            manifest.append({"defectId": defect_id, "page": pno + 1, "file": fname, "width": w, "height": h})

    print(f"Extracted {len(manifest)} defect images")
    return manifest


def load_excel_rows(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.worksheets[0]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column

    def cellval(row, header):
        col = headers.get(header)
        return ws.cell(row=row, column=col).value if col else None

    excel_by_id = {}
    for r in range(2, ws.max_row + 1):
        defect_id = cellval(r, "defect id")
        poly_raw = cellval(r, "polygon data/coordinates")
        if not defect_id or not poly_raw:
            continue
        try:
            points = json.loads(poly_raw)
        except Exception:
            continue
        excel_by_id[str(defect_id)] = {
            "points": points,
            "component": cellval(r, "component"),
            "severity": cellval(r, "severity"),
            "type": cellval(r, "type"),
            "location": cellval(r, "location(m)"),
            "widthMm": cellval(r, "width(mm)"),
            "lengthMm": cellval(r, "length(mm)"),
            "defectStatus": cellval(r, "defect status"),
        }
    return excel_by_id


def measure_marker_bbox(image_path, severity):
    img = Image.open(image_path).convert("RGB")
    arr = np.array(img).astype(np.float32)
    target = SEVERITY_COLORS.get(severity)
    candidates = [target] if target is not None else list(SEVERITY_COLORS.values())

    best = None
    for color in candidates:
        dist = np.sqrt(((arr - color) ** 2).sum(axis=-1))
        mask = dist < COLOR_TOLERANCE
        labeled, n = ndimage.label(mask, structure=np.ones((3, 3)))
        for comp_id in range(1, n + 1):
            ys, xs = np.where(labeled == comp_id)
            if len(xs) < 80:
                continue
            bbox_w = xs.max() - xs.min() + 1
            bbox_h = ys.max() - ys.min() + 1
            fill_ratio = len(xs) / (bbox_w * bbox_h)
            if fill_ratio > 0.4:
                continue
            if best is None or len(xs) > best["pixelCount"]:
                best = {
                    "minX": int(xs.min()), "maxX": int(xs.max()),
                    "minY": int(ys.min()), "maxY": int(ys.max()),
                    "imgW": img.width, "imgH": img.height,
                    "pixelCount": int(len(xs)),
                }
    return best


def cross_reference(excel_by_id, pdf_manifest, pdf_dir):
    pdf_by_id = {m["defectId"]: m for m in pdf_manifest}
    results = []

    for defect_id, excel_row in excel_by_id.items():
        pdf_entry = pdf_by_id.get(defect_id)
        if not pdf_entry:
            continue
        image_path = os.path.join(pdf_dir, pdf_entry["file"])
        measured = measure_marker_bbox(image_path, excel_row.get("severity"))
        if not measured:
            results.append({"defectId": defect_id, **{k: v for k, v in excel_row.items() if k != "points"}, "note": "marker not detected"})
            continue

        sx = NATIVE_W / measured["imgW"]
        sy = NATIVE_H / measured["imgH"]
        m_minX, m_maxX = measured["minX"] * sx, measured["maxX"] * sx
        m_minY, m_maxY = measured["minY"] * sy, measured["maxY"] * sy
        measured_cx, measured_cy = (m_minX + m_maxX) / 2, (m_minY + m_maxY) / 2
        measured_w, measured_h = m_maxX - m_minX, m_maxY - m_minY

        pts = excel_row["points"]
        xs = [p["x"] for p in pts]
        ys = [abs(p["y"]) for p in pts]
        raw_cx, raw_cy = (min(xs) + max(xs)) / 2, (min(ys) + max(ys)) / 2
        raw_w, raw_h = max(xs) - min(xs), max(ys) - min(ys)

        results.append({
            "defectId": defect_id,
            "component": excel_row["component"], "severity": excel_row["severity"],
            "type": excel_row["type"], "location": excel_row["location"],
            "defectStatus": excel_row["defectStatus"],
            "raw_cx": round(raw_cx, 1), "raw_cy": round(raw_cy, 1),
            "raw_w": round(raw_w, 1), "raw_h": round(raw_h, 1),
            "measured_cx": round(measured_cx, 1), "measured_cy": round(measured_cy, 1),
            "measured_w": round(measured_w, 1), "measured_h": round(measured_h, 1),
            "offset_x": round(measured_cx - raw_cx, 1), "offset_y": round(measured_cy - raw_cy, 1),
            "scale_w": round(measured_w / raw_w, 3) if raw_w else None,
            "scale_h": round(measured_h / raw_h, 3) if raw_h else None,
            "match": "NATIVE_OK" if abs(measured_cx - raw_cx) < 80 and abs(measured_cy - raw_cy) < 80 else "MISMATCH",
        })
    return results


def main():
    excel_path, pdf_path, out_dir = sys.argv[1], sys.argv[2], sys.argv[3]
    pdf_dir = os.path.join(out_dir, "pdf_extracted")
    os.makedirs(pdf_dir, exist_ok=True)

    print(f"=== Extracting images from PDF: {pdf_path} ===")
    manifest = extract_pdf_images(pdf_path, pdf_dir)
    with open(os.path.join(pdf_dir, "_pdf_manifest.json"), "w") as f:
        json.dump(manifest, f, indent=2)

    print(f"\n=== Loading Excel: {excel_path} ===")
    excel_by_id = load_excel_rows(excel_path)
    print(f"Loaded {len(excel_by_id)} rows with polygon data")

    print(f"\n=== Cross-referencing ===")
    results = cross_reference(excel_by_id, manifest, pdf_dir)
    with open(os.path.join(out_dir, "_cross_reference.json"), "w") as f:
        json.dump(results, f, indent=2)

    ok = [r for r in results if r.get("match") == "NATIVE_OK"]
    mismatch = [r for r in results if r.get("match") == "MISMATCH"]
    undetected = [r for r in results if "note" in r]
    print(f"\nTotal cross-referenced: {len(results)}")
    print(f"  NATIVE_OK: {len(ok)}")
    print(f"  MISMATCH: {len(mismatch)}")
    print(f"  undetected: {len(undetected)}")

    valid = [r for r in results if "note" not in r and r.get("raw_w", 0) > 5 and r.get("raw_h", 0) > 5]
    if valid:
        scale_w = np.array([r["scale_w"] for r in valid if r["scale_w"]])
        scale_h = np.array([r["scale_h"] for r in valid if r["scale_h"]])
        print(f"\nscale_w: mean={scale_w.mean():.3f} median={np.median(scale_w):.3f}")
        print(f"scale_h: mean={scale_h.mean():.3f} median={np.median(scale_h):.3f}")

    print(f"\nOutput written to: {out_dir}")


if __name__ == "__main__":
    main()
