import json
import openpyxl

EXCEL_FILES = {
    "VSR05-06": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR05-06.xlsx",
    "VSR07-04": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR07-04.xlsx",
    "VSR22-02": r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR22-02.xlsx",
}

date_by_id = {}
for turbine, path in EXCEL_FILES.items():
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column
    id_col = headers.get("defect id")
    date_col = headers.get("inspection date")
    for r in range(2, ws.max_row + 1):
        did = ws.cell(row=r, column=id_col).value
        date = ws.cell(row=r, column=date_col).value
        if did:
            date_by_id[str(did)] = str(date)[:16]

all_rows = []
with open("_cross_reference.json") as f:
    all_rows += json.load(f)
with open("VSR07-04/_cross_reference.json") as f:
    all_rows += json.load(f)
with open("VSR22-02/_cross_reference.json") as f:
    all_rows += json.load(f)

by_date = {}
for r in all_rows:
    if "note" in r or not r.get("scale_h") or r.get("raw_w", 0) < 5:
        continue
    date = date_by_id.get(r["defectId"], "UNKNOWN")
    by_date.setdefault(date, []).append(r)

print("All dates across all 3 turbines:")
for date, rows in sorted(by_date.items()):
    scale_hs = [r["scale_h"] for r in rows]
    avg = sum(scale_hs) / len(scale_hs)
    print(f"  {date}: n={len(rows):3d}  avg_scale_h={avg:.3f}")

# Split into OLD (before 07-30) and NEW (07-30 onward)
NEW_CUTOFF = "2026-07-30"
old_rows = [r for date, rows in by_date.items() for r in rows if date < NEW_CUTOFF]
new_rows = [r for date, rows in by_date.items() for r in rows if date >= NEW_CUTOFF]
print(f"\nOLD format: n={len(old_rows)}")
print(f"NEW format: n={len(new_rows)}")

with open("_new_format_calibration.json", "w") as f:
    json.dump(new_rows, f, indent=2)
with open("_old_format_calibration.json", "w") as f:
    json.dump(old_rows, f, indent=2)
print("\nSaved _new_format_calibration.json and _old_format_calibration.json")
