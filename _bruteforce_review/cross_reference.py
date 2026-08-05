import json
import os
import openpyxl
from PIL import Image
import numpy as np

HERE = os.path.dirname(__file__)
EXCEL_PATH = r"C:\Users\Pedro\Downloads\ATW-2024-0063-2-NAWP-LAGOA DOS VENTOS-VSR05-06.xlsx"
PDF_MANIFEST = os.path.join(HERE, "pdf_extracted", "_pdf_manifest.json")

# 1. Load PDF-extracted image manifest
with open(PDF_MANIFEST) as f:
    pdf_manifest = json.load(f)
pdf_by_id = {m["defectId"]: m for m in pdf_manifest}

# 2. Load Excel rows (raw polygon + metadata) keyed by Defect ID
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.worksheets[0]
headers = {}
for cell in ws[1]:
    if cell.value:
        headers[str(cell.value).strip().lower()] = cell.column

def cellval(row, header):
    col = headers.get(header)
    if not col:
        return None
    return ws.cell(row=row, column=col).value

excel_by_id = {}
for r in range(2, ws.max_row + 1):
    defect_id = cellval(r, "defect id")
    poly_raw = cellval(r, "polygon data/coordinates")
    if not defect_id or not poly_raw:
        continue
    try:
        points = json.loads(poly_raw)
    except Exception:
        continue
    excel_by_id[str(defect_id)] = {
        "points": points,
        "component": cellval(r, "component"),
        "severity": cellval(r, "severity"),
        "type": cellval(r, "type"),
        "location": cellval(r, "location(m)"),
        "widthMm": cellval(r, "width(mm)"),
        "lengthMm": cellval(r, "length(mm)"),
        "defectStatus": cellval(r, "defect status"),
    }

# Known KIN-severity marker colors, sampled directly from extracted PDF images (see find_severity_colors.py)
SEVERITY_COLORS = {
    1: np.array([15, 181, 90]),    # green
    2: np.array([255, 192, 0]),    # gold
    3: np.array([204, 102, 0]),    # dark orange
}
COLOR_TOLERANCE = 25  # euclidean RGB distance

from scipy import ndimage

# 3. For each defect with both a PDF image and Excel polygon, measure the marker bbox.
# Uses connected-component analysis + a "fill ratio" (pixel count / bbox area) filter to
# distinguish the thin polygon OUTLINE from solid-colored blade paint/trim that happens to
# be close to the same severity color (a plain color mask alone grabs the whole green trim
# stripe, since it's near-identical in hue to the severity-1 marker color).
def measure_marker_bbox(image_path, severity):
    img = Image.open(image_path).convert("RGB")
    arr = np.array(img).astype(np.float32)

    target = SEVERITY_COLORS.get(severity)
    candidates = [target] if target is not None else list(SEVERITY_COLORS.values())

    best = None
    for color in candidates:
        dist = np.sqrt(((arr - color) ** 2).sum(axis=-1))
        mask = dist < COLOR_TOLERANCE
        labeled, n = ndimage.label(mask, structure=np.ones((3, 3)))
        for comp_id in range(1, n + 1):
            ys, xs = np.where(labeled == comp_id)
            if len(xs) < 80:
                continue
            bbox_w = xs.max() - xs.min() + 1
            bbox_h = ys.max() - ys.min() + 1
            fill_ratio = len(xs) / (bbox_w * bbox_h)
            if fill_ratio > 0.4:
                continue  # solid-colored area (paint/trim), not a thin outline
            score = len(xs)
            if best is None or score > best["pixelCount"]:
                best = {
                    "minX": int(xs.min()), "maxX": int(xs.max()),
                    "minY": int(ys.min()), "maxY": int(ys.max()),
                    "imgW": img.width, "imgH": img.height,
                    "pixelCount": int(len(xs)),
                }
    return best

NATIVE_W, NATIVE_H = 5568, 4176
results = []

for defect_id, excel_row in excel_by_id.items():
    pdf_entry = pdf_by_id.get(defect_id)
    if not pdf_entry:
        continue
    image_path = os.path.join(HERE, "pdf_extracted", pdf_entry["file"])
    measured = measure_marker_bbox(image_path, excel_row.get("severity"))
    if not measured:
        results.append({"defectId": defect_id, **excel_row, "note": "marker not detected"})
        continue

    # scale measured bbox (from PDF render resolution) to native 5568x4176
    sx = NATIVE_W / measured["imgW"]
    sy = NATIVE_H / measured["imgH"]
    measured_native = {
        "minX": measured["minX"] * sx, "maxX": measured["maxX"] * sx,
        "minY": measured["minY"] * sy, "maxY": measured["maxY"] * sy,
    }
    measured_cx = (measured_native["minX"] + measured_native["maxX"]) / 2
    measured_cy = (measured_native["minY"] + measured_native["maxY"]) / 2
    measured_w = measured_native["maxX"] - measured_native["minX"]
    measured_h = measured_native["maxY"] - measured_native["minY"]

    # raw excel polygon bbox, AS-IS (x, abs(y)), no transform
    pts = excel_row["points"]
    xs = [p["x"] for p in pts]
    ys = [abs(p["y"]) for p in pts]
    raw_cx = (min(xs) + max(xs)) / 2
    raw_cy = (min(ys) + max(ys)) / 2
    raw_w = max(xs) - min(xs)
    raw_h = max(ys) - min(ys)

    results.append({
        "defectId": defect_id,
        "component": excel_row["component"],
        "severity": excel_row["severity"],
        "type": excel_row["type"],
        "location": excel_row["location"],
        "defectStatus": excel_row["defectStatus"],
        "raw_cx": round(raw_cx, 1), "raw_cy": round(raw_cy, 1),
        "raw_w": round(raw_w, 1), "raw_h": round(raw_h, 1),
        "measured_cx": round(measured_cx, 1), "measured_cy": round(measured_cy, 1),
        "measured_w": round(measured_w, 1), "measured_h": round(measured_h, 1),
        "offset_x": round(measured_cx - raw_cx, 1),
        "offset_y": round(measured_cy - raw_cy, 1),
        "scale_w": round(measured_w / raw_w, 3) if raw_w else None,
        "scale_h": round(measured_h / raw_h, 3) if raw_h else None,
        "match": "NATIVE_OK" if abs(measured_cx - raw_cx) < 80 and abs(measured_cy - raw_cy) < 80 else "MISMATCH",
    })

with open(os.path.join(HERE, "_cross_reference.json"), "w") as f:
    json.dump(results, f, indent=2)

ok = [r for r in results if r.get("match") == "NATIVE_OK"]
mismatch = [r for r in results if r.get("match") == "MISMATCH"]
undetected = [r for r in results if "note" in r]

print(f"Total cross-referenced: {len(results)}")
print(f"  NATIVE_OK (raw coordinate already matches, within 80px): {len(ok)}")
print(f"  MISMATCH (raw coordinate does NOT match measured position): {len(mismatch)}")
print(f"  undetected marker: {len(undetected)}")

print("\n=== MISMATCHES ===")
for r in mismatch:
    print(f"  {r['defectId']:16s} comp={str(r['component']):14s} sev={r['severity']} loc={r['location']:>6} "
          f"offset=({r['offset_x']:+7.1f},{r['offset_y']:+7.1f}) scale=({r['scale_w']},{r['scale_h']}) status={r['defectStatus']}")
